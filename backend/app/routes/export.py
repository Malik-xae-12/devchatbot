from __future__ import annotations

import io
import re

from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.services.export_cache import export_cache

router = APIRouter(prefix="/api/export", tags=["export"])

HEADER_FILL = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _safe_filename(question: str) -> str:
    slug = re.sub(r"[^a-zA-Z0-9]+", "_", question.strip()).strip("_").lower()
    return (slug[:40] or "query_results") + ".xlsx"


@router.get("/{export_id}")
async def download_export(export_id: str) -> StreamingResponse:
    entry = export_cache.get(export_id)
    if entry is None:
        raise HTTPException(
            status_code=404,
            detail="This download has expired or was already used. Ask the question again to regenerate it.",
        )

    rows = entry.rows
    columns = list(rows[0].keys()) if rows else []

    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    ws.append(columns)
    for col_idx in range(1, len(columns) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.freeze_panes = "A2"

    for row in rows:
        ws.append([row.get(col) for col in columns])

    for col_idx, col_name in enumerate(columns, start=1):
        longest = max([len(col_name)] + [len(str(row.get(col_name, ""))) for row in rows])
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(longest + 2, 10), 60)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = _safe_filename(entry.question)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
